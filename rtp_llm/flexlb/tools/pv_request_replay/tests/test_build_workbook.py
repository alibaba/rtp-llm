from __future__ import annotations

import json
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


TOOL_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(TOOL_DIR))
import build_workbook as workbook_module  # noqa: E402


SHANGHAI = ZoneInfo("Asia/Shanghai")


def epoch_ms(hour: int, minute: int, second: int = 0) -> int:
    return int(
        datetime(2026, 8, 11, hour, minute, second, tzinfo=SHANGHAI).timestamp()
        * 1000
    )


def pv_line(log_time: str, record: dict) -> str:
    return (
        f"{log_time} INFO test pvLogger - "
        + json.dumps(record, separators=(",", ":"))
        + "\n"
    )


def route_record(request_id: str, request_time_ms: int, worker: str) -> dict:
    return {
        "requestId": request_id,
        "requestTimeMs": request_time_ms,
        "totalUs": 1234,
        "success": True,
        "inputIdsCount": 1000,
        "selectionReasons": {"PREFILL": "SHORTEST_TTFT"},
        "response": {
            "code": 200,
            "server_status": [
                {"role": "PREFILL", "server_ip": worker, "code": 200}
            ],
        },
        "shortestTtftDecisions": [
            {
                "role": "PREFILL",
                "decisionTimeMs": request_time_ms,
                "routingAttempt": 1,
                "workers": [
                    {
                        "ip": worker,
                        "port": 8001,
                        "selected": True,
                        "estimatedTtftRank": 1,
                        "requestHitRatePct": 80,
                        "requestUncachedTokens": 200,
                    }
                ],
            }
        ],
    }


def cache_record(request_id: str, actual_hit: int) -> dict:
    return {
        "event": "cache_hit_comparison",
        "requestId": request_id,
        "inputTokens": 1000,
        "kvcm": {"hit": 800},
        "actual": {"hit": actual_hit},
        "state": "running",
    }


def status_record(request_id: str, worker: str, request_time_ms: int) -> dict:
    enqueue = request_time_ms + 100
    return {
        "event": "prefill_worker_status",
        "requestId": request_id,
        "workerIp": worker,
        "inputQueueEnqueueTimeMs": enqueue,
        "inputQueueDrainTimeMs": enqueue + 10,
        "firstTokenTimeMs": enqueue + 1010,
        "inputQueueWaitMs": 10,
        "schedulerWaitMs": 100,
        "remoteKvWaitMs": 200,
        "schedulerToRunningMs": 300,
        "runningToFirstTokenMs": 700,
        "hbmLocalMatchTokens": 300,
        "remoteKvAddedMatchTokens": 500,
        "prefillStepCount": 1,
        "firstPrefillStepId": 7,
        "lastPrefillStepId": 7,
    }


class BuildWorkbookTest(unittest.TestCase):
    def test_joins_by_instance_and_filters_only_route_request_time(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source_a = root / "raw" / "flexlb-a" / "pv.log.snapshot"
            source_b = root / "raw" / "flexlb-b" / "pv.log.snapshot"
            source_a.parent.mkdir(parents=True)
            source_b.parent.mkdir(parents=True)

            request_a_ms = epoch_ms(1, 45)
            request_b_ms = epoch_ms(1, 46)
            outside_ms = epoch_ms(2, 0)
            source_a.write_text(
                "".join(
                    [
                        pv_line(
                            "2026-08-11 01:45:00.010",
                            route_record("same-request", request_a_ms, "10.0.0.1"),
                        ),
                        # Terminal records deliberately arrive after the report
                        # window.  They must still join to the in-window route.
                        pv_line(
                            "2026-08-11 02:03:00.000",
                            cache_record("same-request", 750),
                        ),
                        pv_line(
                            "2026-08-11 02:03:01.000",
                            status_record("same-request", "10.0.0.1", request_a_ms),
                        ),
                        pv_line(
                            "2026-08-11 01:59:59.000",
                            route_record("end-exclusive", outside_ms, "10.0.0.9"),
                        ),
                        pv_line(
                            "2026-08-11 02:04:00.000",
                            status_record("end-exclusive", "10.0.0.9", outside_ms),
                        ),
                    ]
                ),
                encoding="utf-8",
            )
            source_b.write_text(
                "".join(
                    [
                        pv_line(
                            "2026-08-11 01:46:00.010",
                            route_record("same-request", request_b_ms, "10.0.0.2"),
                        ),
                        pv_line(
                            "2026-08-11 02:02:00.000",
                            cache_record("same-request", 900),
                        ),
                        pv_line(
                            "2026-08-11 02:02:01.000",
                            status_record("same-request", "10.0.0.2", request_b_ms),
                        ),
                    ]
                ),
                encoding="utf-8",
            )

            destination = root / "analysis.xlsx"
            summary = workbook_module.build_workbook(
                sources=[source_a, source_b],
                destination=destination,
                start=datetime(2026, 8, 11, 1, 40, tzinfo=SHANGHAI),
                end=datetime(2026, 8, 11, 2, 0, tzinfo=SHANGHAI),
            )

            self.assertEqual(summary["source_count"], 2)
            self.assertEqual(summary["instance_count"], 2)
            self.assertEqual(summary["request_count"], 2)
            self.assertEqual(summary["complete_request_count"], 2)
            self.assertEqual(summary["event_counts"]["route_outside_window"], 1)

            workbook = load_workbook(destination, read_only=True, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Requests",
                    "P99 Focus",
                    "Decision Snapshot Top5",
                    "Host Summary",
                    "Data Scope",
                ],
            )
            worksheet = workbook["Requests"]
            headers = [cell.value for cell in worksheet[4]]
            request_id_col = headers.index("request_id")
            instance_col = headers.index("flexlb_instance")
            request_rows = [
                row
                for row in worksheet.iter_rows(min_row=5, values_only=True)
                if row[request_id_col] == "same-request"
            ]
            self.assertEqual(len(request_rows), 2)
            self.assertEqual(
                {row[instance_col] for row in request_rows}, {"flexlb-a", "flexlb-b"}
            )

            snapshot_sheet = workbook["Decision Snapshot Top5"]
            snapshot_headers = [cell.value for cell in snapshot_sheet[4]]
            guard_col = snapshot_headers.index("outstanding_guard_eligible")
            guard_values = [
                row[guard_col]
                for row in snapshot_sheet.iter_rows(min_row=5, values_only=True)
            ]
            self.assertEqual(guard_values, [None, None])

    def test_prefill_extractors_do_not_fall_back_to_decode_role(self) -> None:
        decode_only = {
            "response": {"server_status": [{"role": "DECODE", "server_ip": "decode"}]},
            "cacheMatchSelections": [{"role": "DECODE", "selectedIp": "decode"}],
            "shortestTtftDecisions": [{"role": "DECODE", "workers": [{"ip": "decode"}]}],
        }
        self.assertEqual(workbook_module.get_prefill_server_status(decode_only), {})
        self.assertEqual(workbook_module.get_route_cache_selection(decode_only), {})
        self.assertEqual(workbook_module.get_prefill_decision(decode_only), {})

        legacy = {
            "response": {"server_status": [{"server_ip": "legacy"}]},
            "cacheMatchSelections": [{"selectedIp": "legacy"}],
            "shortestTtftDecisions": [{"workers": [{"ip": "legacy"}]}],
        }
        self.assertEqual(
            workbook_module.get_prefill_server_status(legacy)["server_ip"], "legacy"
        )
        self.assertEqual(
            workbook_module.get_route_cache_selection(legacy)["selectedIp"], "legacy"
        )
        self.assertEqual(
            workbook_module.get_prefill_decision(legacy)["workers"][0]["ip"],
            "legacy",
        )

        ambiguous_legacy = {
            "response": {"server_status": [{"server_ip": "one"}, {"server_ip": "two"}]},
            "cacheMatchSelections": [{"selectedIp": "one"}, {"selectedIp": "two"}],
            "shortestTtftDecisions": [{"workers": []}, {"workers": []}],
        }
        self.assertEqual(workbook_module.get_prefill_server_status(ambiguous_legacy), {})
        self.assertEqual(workbook_module.get_route_cache_selection(ambiguous_legacy), {})
        self.assertEqual(workbook_module.get_prefill_decision(ambiguous_legacy), {})

    def test_optional_boolean_keeps_unknown_distinct_from_false(self) -> None:
        self.assertEqual(workbook_module.yes_no_unknown(None), "")
        self.assertEqual(workbook_module.yes_no_unknown(False), "NO")
        self.assertEqual(workbook_module.yes_no_unknown(True), "YES")

    def test_discovers_collector_manifest_and_preserves_instance(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            snapshot = root / "raw" / "instance-from-parent" / "pv.log.snapshot"
            snapshot.parent.mkdir(parents=True)
            snapshot.write_text("", encoding="utf-8")
            manifest = root / "collect_manifest.json"
            manifest.write_text(
                json.dumps(
                    {
                        "snapshots": [
                            {"instance": "instance-from-manifest", "path": str(snapshot)}
                        ]
                    }
                ),
                encoding="utf-8",
            )

            direct = workbook_module.discover_sources(snapshot)
            manifested = workbook_module.discover_sources(root)
            self.assertEqual(direct[0].instance, "instance-from-parent")
            self.assertEqual(manifested[0].instance, "instance-from-manifest")

    def test_empty_route_selection_fails_before_creating_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "pv.log"
            destination = root / "analysis.xlsx"
            source.write_text("unrelated log line\n", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "No routing PV records matched"):
                workbook_module.build_workbook(source, destination)
            self.assertFalse(destination.exists())


if __name__ == "__main__":
    unittest.main()
