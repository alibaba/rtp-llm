#!/usr/bin/env python3
"""Build a self-contained FlexLB Prefill request replay page.

The input workbook is treated as read-only. The generated HTML embeds the
compact replay data and all UI assets, so it can be shared and opened directly
with ``file://`` without a web service or CDN.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


SHANGHAI = ZoneInfo("Asia/Shanghai")
REQUESTS_SHEET = "Requests"
CANDIDATES_SHEET = "Decision Snapshot Top5"
HEADER_ROW = 4
DATA_PLACEHOLDER = "__REPLAY_DATA__"


def replay_key(row: dict[str, Any]) -> str:
    """Keep equal request IDs from different FlexLB instances independent."""
    request_id = str(row.get("request_id") or "")
    instance = str(row.get("flexlb_instance") or "")
    return f"{instance}::{request_id}" if instance else request_id


def cell_number(value: Any) -> int | float | None:
    """Return an Excel value as a compact JSON number when possible."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)

    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        parsed = float(text)
        return int(parsed) if parsed.is_integer() else parsed
    except ValueError:
        return None


def timestamp_ms(value: Any) -> int | None:
    """Convert workbook timestamps to Unix milliseconds in Asia/Shanghai."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        try:
            parsed = datetime.fromisoformat(str(value).strip())
        except ValueError:
            return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=SHANGHAI)
    return int(parsed.timestamp() * 1000)


def cell_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "yes", "1", "y"}


def read_table(worksheet: Any, header_row: int = HEADER_ROW) -> list[dict[str, Any]]:
    headers = [
        cell.value
        for cell in next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row)
        )
    ]
    records: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        record = {
            header: value
            for header, value in zip(headers, row)
            if header is not None and header != ""
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def compact_request(row: dict[str, Any]) -> dict[str, Any] | None:
    request_id = row.get("request_id")
    if not request_id or request_id == "section":
        return None
    route = timestamp_ms(row.get("route_log_time (decision)"))
    if route is None:
        return None

    number = lambda name: cell_number(row.get(name))
    return {
        "id": replay_key(row),
        "requestId": str(request_id),
        "flexlbInstance": str(row.get("flexlb_instance") or ""),
        "host": str(row.get("prefill_host") or "unknown"),
        "route": route,
        "enqueue": timestamp_ms(row.get("input_queue_enqueue_time")),
        "drain": timestamp_ms(row.get("input_queue_drain_time")),
        "firstToken": timestamp_ms(row.get("first_token_time")),
        "cacheEvent": timestamp_ms(row.get("cache_event_time")),
        "reason": row.get("selection_reason") or "UNKNOWN",
        "snapshotStatus": row.get("decision_snapshot_status") or "",
        "routingAttempt": number("decision_routing_attempt"),
        "cacheLeader": row.get("decision cache leader") or "",
        "shortestTtft": row.get("decision shortest TTFT") or "",
        "cacheLead": number("decision cache lead tokens"),
        "extraWork": number("decision extra work tokens"),
        "maxExtraWork": number("decision max extra work tokens"),
        "outstandingThreshold": number("decision outstanding threshold"),
        "selectedRank": number("selected snapshot TTFT rank"),
        "selectedSnapshotHitRate": number("selected snapshot hit rate"),
        "selectedSnapshotUncache": number("selected snapshot request uncache"),
        "selectedQueueWork": number("selected snapshot queue work"),
        "selectedEstimatedTtft": number("selected snapshot estimated TTFT"),
        "selectedOutstanding": number("selected outstanding uncache"),
        "selectedOutstandingAfter": number("selected outstanding after request"),
        "engineWaiting": number("selected engine waiting uncache"),
        "engineRunningRemaining": number("selected engine RUNNING remaining"),
        "responseCode": number("route_response_code"),
        "telemetryStatus": row.get("telemetry_status") or "",
        "engineTtft": number("prefill_engine_ttft_ms"),
        "percentile": row.get("prefill_ttft_percentile") or "",
        "routeToFirst": number("route_to_first_token_ms"),
        "routeToEnqueue": number("route_to_engine_enqueue_ms"),
        "dominantPhase": row.get("dominant_observed_phase") or "",
        "inputQueueWait": number("input_queue_wait_ms"),
        "schedulerWait": number("scheduler_wait_ms"),
        "remoteKvWait": number("remote_kv_wait_ms (scheduler subset)"),
        "schedulerToRunning": number("scheduler_to_running_ms"),
        "runningToFirst": number("running_to_first_token_ms"),
        "hbmLocal": number("hbm_local_match_tokens"),
        "remoteKvAdded": number("remote_kv_added_match_tokens"),
        "stepCount": number("prefill_step_count"),
        "firstStep": number("first_prefill_step_id"),
        "lastStep": number("last_prefill_step_id"),
        "chunkMin": number("prefill_nonfinal_chunk_tokens_min"),
        "chunkMax": number("prefill_nonfinal_chunk_tokens_max"),
        "inputTokens": number("input_tokens"),
        "uncache": number("uncache_tokens"),
        "actualHitRate": number("actual_hit_rate_pct"),
        "predictedHitRate": number("predicted_hit_rate_pct"),
        "hitDeltaPp": number("actual_minus_predicted_pp"),
        "predictedHit": number("predicted_hit_tokens"),
        "actualHit": number("actual_hit_tokens"),
        "cacheState": row.get("cache_state") or "",
        "workerStatusEvent": timestamp_ms(row.get("worker_status_event_time")),
    }


def compact_candidate(row: dict[str, Any]) -> dict[str, Any] | None:
    if not row.get("request_id"):
        return None
    number = lambda name: cell_number(row.get(name))
    eligible_value = row.get("outstanding_guard_eligible")
    return {
        "rank": number("candidate_rank_by_estimated_TTFT"),
        "host": str(row.get("candidate_ip") or "unknown"),
        "port": number("candidate_port"),
        "selected": cell_bool(row.get("selected")),
        "cacheLeader": cell_bool(row.get("cache_leader")),
        "shortest": cell_bool(row.get("shortest_TTFT")),
        "eligible": (
            None
            if eligible_value in (None, "")
            else cell_bool(eligible_value)
        ),
        "hitRate": number("request_hit_rate_pct"),
        "hitTokens": number("request_hit_cache_tokens"),
        "uncache": number("request_uncache_tokens"),
        "prefillWork": number("request_prefill_work"),
        "queueWork": number("queue_work_before_route"),
        "estimatedTtft": number("estimated_TTFT_work"),
        "outstanding": number("outstanding_uncache_before"),
        "outstandingAfter": number("outstanding_uncache_after"),
        "inTransitWaiting": number("in_transit_waiting_uncache"),
        "trackedRunningRemaining": number("tracked_RUNNING_remaining"),
        "engineWaiting": number("engine_WAITING_uncache"),
        "engineRunningRemaining": number("engine_RUNNING_remaining"),
        "availableKv": number("available_KV_cache_tokens"),
        "usedKv": number("used_KV_cache_tokens"),
        "statusAgeUs": number("status_age_us"),
    }


def _build_replay(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        missing = [
            name
            for name in (REQUESTS_SHEET, CANDIDATES_SHEET)
            if name not in workbook.sheetnames
        ]
        if missing:
            raise ValueError(f"Workbook is missing sheet(s): {', '.join(missing)}")

        requests = [
            item
            for row in read_table(workbook[REQUESTS_SHEET])
            if (item := compact_request(row))
        ]
        if not requests:
            raise ValueError("Workbook has no replayable request rows")
        requests.sort(key=lambda item: item["route"])

        candidates: dict[str, list[dict[str, Any]]] = {}
        for row in read_table(workbook[CANDIDATES_SHEET]):
            item = compact_candidate(row)
            if item:
                candidates.setdefault(replay_key(row), []).append(item)
        for values in candidates.values():
            values.sort(key=lambda item: (item["rank"] or 10**9, item["host"]))
    finally:
        workbook.close()

    hosts = sorted({request["host"] for request in requests})
    terminal_times = [
        request["firstToken"] or request["route"] for request in requests
    ]
    return {
        "meta": {
            "source": input_path.name,
            "requestCount": len(requests),
            "hostCount": len(hosts),
            "candidateCount": sum(len(values) for values in candidates.values()),
            "start": requests[0]["route"],
            "end": max(terminal_times),
            "timezone": "Asia/Shanghai",
            "notice": (
                "决策候选为 PV 中的真实 Top5/关键候选快照；请求生命周期和 "
                "step 进度依据观测边界展示。"
            ),
        },
        "hosts": hosts,
        "requests": requests,
        "candidates": candidates,
    }


def build_html(
    input_path: Path, template_path: Path, output_path: Path
) -> dict[str, Any]:
    """Build standalone replay HTML and return a compact build summary."""
    input_path = Path(input_path)
    template_path = Path(template_path)
    output_path = Path(output_path)

    replay = _build_replay(input_path)
    template = template_path.read_text(encoding="utf-8")
    placeholder_count = template.count(DATA_PLACEHOLDER)
    if placeholder_count != 1:
        raise ValueError(
            f"Template must contain exactly one {DATA_PLACEHOLDER} placeholder; "
            f"found {placeholder_count}"
        )

    payload = json.dumps(replay, ensure_ascii=False, separators=(",", ":"))
    payload = (
        payload.replace("<", "\\u003c")
        .replace("\u2028", "\\u2028")
        .replace("\u2029", "\\u2029")
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        template.replace(DATA_PLACEHOLDER, payload), encoding="utf-8"
    )

    meta = replay["meta"]
    return {
        "input": str(input_path),
        "output": str(output_path),
        "request_count": meta["requestCount"],
        "host_count": meta["hostCount"],
        "candidate_count": meta["candidateCount"],
        "start_ms": meta["start"],
        "end_ms": meta["end"],
    }
